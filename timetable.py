import random
import openpyxl
import os
import sys
import math

from openpyxl.styles.fonts import Font
from openpyxl.styles import PatternFill


wb = openpyxl.load_workbook("order_sheet.xlsx")
sheet = wb["Sheet1"]

#200人分の名前スペースを用意
name = [0 for num in range(200)]

#人数をカウントする
name_number = 0

#試合数
male_game_number = sheet.cell(3,7).value
female_game_number = sheet.cell(4,7).value
game_number = male_game_number + female_game_number

#サイクル数
cycle_number = sheet.cell(5,7).value

#1サイクルの限界試合数
game_number_percycle = round(game_number / cycle_number) + 1

#男子限界サイクル数
male_game_number_percycle = round(male_game_number / cycle_number) + 1

#コート数
court_number = sheet.cell(6,7).value


#昼休憩
afternoon = 4
if sheet.cell(8,7).value == "あり":
   lunch_break = 1
   lunch = "昼休憩あり"
   if sheet.cell(9,7).value != None:
      afternoon = sheet.cell(9,7).value -1
   afternoon_time = "午後"+str(afternoon+1)+"サイから"
   half_time = afternoon
   
else:
   lunch_break = 0
   lunch = "昼休憩なし"
   afternoon_time = ""
   afternoon = 0
   half_time = round(cycle_number/2)



#3連サイを消すか否か
if sheet.cell(3,17).value == "なし":
   three_check = 1
else:
   three_check = 0
   
#名前スペースに名前を入れていく
for colum in range(3,game_number*2+3):
   for row in range(3,5):
      if sheet.cell(colum,row).value not in name:
         name[name_number] = sheet.cell(colum,row).value
         name_number = name_number + 1

#各選手が何回入っているか確認
player_count = [0 for num in range(name_number)]
for colum in range(3,game_number*2+3):
    for row in range(3,5):
        player_count[name.index(sheet.cell(colum,row).value)] += 1

#各試合に誰が入るかを番号で入力
player_g = [[name.index(sheet.cell(num*2+3,3).value),name.index(sheet.cell(num*2+4,3).value),name.index(sheet.cell(num*2+3,4).value),name.index(sheet.cell(num*2+4,4).value)] for num in range(game_number)]


#1つの入れ子の中に1サイクルの試合を全部入れる
player_c_default = [[] for num in range(cycle_number)]

 
#あるサイクルに試合をする4人がどれくらい入っているかの変数
cycle_people = [0 for num in range(cycle_number+2)]

#あるサイクルに入る確率とあるサイクルに入った場合の連サイ数
rate = [0 for num in range(cycle_number)]
cycle_continuous = [0 for num in range(cycle_number)]

#n番の選手はどのサイクルに入ることができるか
number_of_cycle = [num for num in range(cycle_number)]

p_set = [0 for num in range(name_number)]


#試合係
match_engagement = [num for num2 in range(11,17) for num in range(name_number) if name[num] == sheet.cell(num2,7).value]

match_engagement_name = [name[num] for num in match_engagement]

shomu_chief = [num for num in range(name_number) if name[num] == sheet.cell(11,7).value]

shiai_chief = [num for num in range(name_number) if name[num] == sheet.cell(12,7).value]

#会計
treasurer = [num for num2 in range(18,23) for num in range(name_number) if name[num] == sheet.cell(num2,7).value]

#先輩
senior = [num for num3 in range(13,15) for num2 in range(19,51) for num in range(name_number) if name[num] == sheet.cell(num2,num3).value]

#班長入場に出る人
leader_admission = [num for num3 in range(13,15) for num2 in range(3,18) for num in range(name_number) if name[num] == sheet.cell(num2,num3).value]

#班長戦
leader_game = [(sheet.cell(num,10).value - 1) for num in range(3,9) if sheet.cell(num,10).value != None]

#入れたくないサイクル
cycle_not_specified = [[],[]]
cycle_not_specified[0] = [num for num2 in range(13,51) for num in range(name_number) if name[num] == sheet.cell(num2,9).value]
cycle_not_specified[1] = [(sheet.cell(num,10).value - 1) for num in range(13,51) if sheet.cell(num,10).value != None]
not_specified_number = len(cycle_not_specified[0])
if len(cycle_not_specified[0]) != len(cycle_not_specified[1]):
   print("入れたくないサイクルが間違っています")
   sys.exit()

#サイクル指定
cycle_specified = [[],[]]
cycle_specified[0] = [(sheet.cell(num,6).value - 1) for num in range(26,51) if sheet.cell(num,6).value != None]
cycle_specified[1] = [(sheet.cell(num,7).value - 1) for num in range(26,51) if sheet.cell(num,7).value != None]
if len(cycle_specified[0]) != len(cycle_specified[1]):
   print("入れたいサイクルが間違ってます")
   sys.exit()
   
#サイクル指定と班長戦のサイクルを除外
random_game = set(range(game_number)).difference(set(cycle_specified[0]))
random_game = random_game.difference(set(leader_game))

#班名入力
try:
   sheet0 = wb["Sheet2"]
   if sheet0.cell(1,1).value != None:
      kainai = 1

      #班名読み込み
      group_name = [sheet0.cell(1,num+1).value for num in range(6)]

      #各班に振り分け
      group_structure = [[],[],[],[],[],[]]
      for num in range(6):
         group_structure[num] = [name.index(sheet0.cell(num2,num+1).value) for num2 in range(2,51) if sheet0.cell(num2,num+1).value in name]

   else:
      kainai = 0
        
except KeyError:
   kainai = 0


#上から10個のタイムテーブルを保存
best_timetable = []
best_continuous_cycle = []
#best_enter_cycle = []
for num in range(10):
   best_timetable.append([])
   #best_enter_cycle.append([])
   for nume in range(cycle_number):
      best_timetable[num].append([])        
   best_continuous_cycle.append(100)

#合計連サイ数
sum_continuous_cycle = 0

#没になったタイムテーブルの数
ng_number = 0

#ループの回数
print("試行回数を入力して下さい")
loop_number = int(input())

print("タイムテーブル作成開始")

########タイムテーブル作成開始########
for loop in range(loop_number):

   #進捗を表示
   if loop_number <= 10000:
      progress_rate = loop*10/loop_number
      if progress_rate.is_integer():
         print(str(int(progress_rate*10))+"%終了")
   elif loop_number <= 30000:
      progress_rate = loop*20/loop_number
      if progress_rate.is_integer():
         print(str(int(progress_rate*5))+"%終了")
   else:
      progress_rate = loop*100/loop_number
      if progress_rate.is_integer():
         print(str(int(progress_rate))+"%終了")
      

   #班長戦の数をリセット
   leader_num = 0

   #NGタイムテーブルリセット
   ng_timetable = 0

   #男子が少ないサイクルがないかチェックリセット
   male_check = 0
   
   #はいった試合を一度リセット
   player_c = [[] for num in range(cycle_number)]
   

   #n番の選手はどのサイクルに入ることができるか
   p_set = [set(number_of_cycle) for num in range(name_number)]
      #セットをそのまま代入すると不具合が起きるのでリストから変換
      
   #どのサイクルに何枠残ってるかを決める
   b_list = [num for num in range(cycle_number) for num2 in range(game_number_percycle)]
 
   #連サイの数
   continuous_cycle = 0

   #各サイクルの試合係の数
   match_engagement_number = [0 for num in range(cycle_number)]
   
   #班長戦のサイクルをリセット
   leader_game_cycle = []

   #男→女の順番で
   for sex in range(2):
      
      ###サイクルが指定された試合を決定###
      for n in range(len(cycle_specified[0])):

         if sex == 0 and cycle_specified[0][n] >= male_game_number:
            continue
         if sex == 1 and cycle_specified[0][n] < male_game_number:
            continue
         
         #試合に入ってる人を確認
         player=[player_g[cycle_specified[0][n]][num] for num in range(4)]
         
         dec_cycle = cycle_specified[1][n]

         b_list.remove(dec_cycle)
                
         #まだ選手が入っていないサイクルから消す
         for num in range(4):
            p_set[player[num]].remove(dec_cycle)
            
         #サイクルに選手を追加
         player_c[dec_cycle].append(player_g[cycle_specified[0][n]])

         #試合係が入ったサイクルをカウント
         for num in range(4):
            if player[num] in match_engagement:
               match_engagement_number[dec_cycle] += 1

         #班長戦が入ったサイクルを記録
         if cycle_specified[0][n] in leader_game:
            leader_game_cycle.append(dec_cycle)   
         
      ###班長戦を決定###
      for n in leader_game:

         if sex == 1:
            break
         
         #班長戦の数をカウント
         leader_num += 1

         #試合に入ってる人を確認
         player=[player_g[n][num] for num in range(4)]
         
         #試合係の人数をカウント
         match_engagement_exist = 0
         for num in range(4):
            if player[num] in match_engagement:
               match_engagement_exist += 1

         #まだ入ってないサイクル
         remain_cycle = [p_set[player[num]] for num in range(4)]

         #4人全員が入れるサイクル
         for num in range(3):
            remain_cycle[num+1] = remain_cycle[num].intersection(remain_cycle[num+1])

         #枠が残ったサイクル
         available_cycle = set(b_list)

         #4人全員が入れて枠が残ったサイクル
         remain_cycle_all = available_cycle.intersection(remain_cycle[3])

         if len(remain_cycle_all) == 0:
            #入れるサイクルがない場合は枠がないサイクルから選ぶ
            remain_cycle_all = remain_cycle[3]    
            not_remain = 1
         else:
            not_remain = 0
              
         #どのサイクルに何人入っているか
         cycle_people = [0 for num in range(cycle_number+2)]
         #-1を定義できないので…

         #今までのどこのサイクルに入ってたか
         for num in range(4):
            for num2 in range(cycle_number):
               if num2 not in p_set[player[num]]:
                  if player[num] in senior:
                     #先輩は連サイなるべくなくす
                     cycle_people[num2+1] = cycle_people[num2+1] + 3
                  else:               
                     cycle_people[num2+1] = cycle_people[num2+1] + 1

         #各サイクルに入る確率と連サイ数をリセット
         rate = [0 for num in range(cycle_number)]
         cycle_continuous = [0 for num in range(cycle_number)]

         #各サイクルに入る確率を決定
         for num in remain_cycle_all:
            #昼休憩がある場合は4,5連サイなし
            if num == afternoon -1 and lunch_break == 1:
               cycle_continuous[num] = cycle_people[num]
            elif num == afternoon and lunch_break == 1:
               cycle_continuous[num] = cycle_people[num+2]
            else:
               cycle_continuous[num] = cycle_people[num] + cycle_people[num+2]
            #2520は1-9の最小公倍数,1人3試合の場合連サイ数は最大8
            rate[num] = pow(50,20) / pow((cycle_continuous[num] + 1),5)       
            #試合係が3人より多くなるの場合はrate=0
            if match_engagement_exist > 0 and match_engagement_number[num] + match_engagement_exist > 3:
               rate[num] = 0
            #班長戦がの入る枠がない場合は1サイクルに２つ
            if leader_num <= cycle_number-2:
               #班長戦がすでに入っていたらrate=0
               if num in leader_game_cycle:
                  rate[num] = 0
            #入れてほしくないサイクル
            for num2 in range(not_specified_number):
               for num3 in range(4):
                  if player[num3] == cycle_not_specified[0][num2] and num ==  cycle_not_specified[1][num2]:
                     rate[num] = 0

         #班長戦は最初と最後には入らない
         rate[0] = 0
         rate[cycle_number-1] = 0
         
         #昼休憩ありの場合は会計は昼前、昼後には入れない。試合チーフは昼前には入れない
         if lunch_break == 1:
            for num in range(4):
               if player[num] in treasurer:
                  rate[afternoon -1] = 0
                  rate[afternoon] = 0
               if player[num] in shiai_chief:
                  rate[afternoon -1] = 0
         
         #各サイクルの確率を足し合わせる      
         sum_rate = 0
         for num in range(cycle_number):
            sum_rate = sum_rate + rate[num]

                    
         if len(remain_cycle_all) != 0 and sum_rate != 0:

            #どのサイクルに入るかを乱数で決定
            random_number = random.randrange(sum_rate)
            #下のfor文に使うためにリセット 
            sum_rate = 0
              
            for num in range(cycle_number):
               if sum_rate <= random_number and random_number < sum_rate + rate[num]:
                  dec_cycle = num
                  continuous_cycle += cycle_continuous[num]
               sum_rate = sum_rate + rate[num]
                 
            #枠が残っている場合だけ消す
            if not_remain == 0:
               b_list.remove(dec_cycle)
                
            #まだ選手が入っていないサイクルから消す
            for num in range(4):
               p_set[player[num]].remove(dec_cycle)
                  
            #サイクルに選手を追加
            player_c[dec_cycle].append(player_g[n])

            #試合係が入ったらカウント
            match_engagement_number[dec_cycle] += match_engagement_exist

            #班長戦が入ったサイクルを記録
            if n in leader_game:
               leader_game_cycle.append(dec_cycle)
               
         else:
             ng_timetable = 1
      

      ###サイクルが指定されてない試合を決定###
      for n in random_game:

         if sex == 0 and n >= male_game_number:
            continue
         if sex == 1 and n < male_game_number:
            continue

         #男子の試合数が少ないサイクルがある場合は排除
         if n >= male_game_number and male_check == 0:
            length_cycle = [len(player_c[num]) for num in range(cycle_number)]
            male_check = 1
            if min(length_cycle) <= male_game_number_percycle-3:
               ng_timetable = 1
               break
      
         #試合に入ってる人を確認
         player=[player_g[n][num] for num in range(4)]

         #試合係の人数をカウント
         match_engagement_exist = 0
         for num in range(4):
            if player[num] in match_engagement:
               match_engagement_exist += 1

         #まだ入ってないサイクル
         remain_cycle = [p_set[player[num]] for num in range(4)]
         #入ったサイクル
         enter_cycle = [set(number_of_cycle).difference(remain_cycle[num]) for num in range(4)]

         #4人全員が入れるサイクル
         for num in range(3):
            remain_cycle[num+1] = remain_cycle[num].intersection(remain_cycle[num+1])
         #枠が残ったサイクル
         available_cycle = set(b_list)
         #4人全員が入れて枠が残ったサイクル
         remain_cycle_all = available_cycle.intersection(remain_cycle[3])

         if len(remain_cycle_all) == 0:
            #入れるサイクルがない場合はタイムテーブル放棄
            ng_timetable = 1
            break
              
         #どのサイクルに何人入っているか
         cycle_people = [0 for num in range(cycle_number+2)]
         #-1を定義できないので…

         #今までのどこのサイクルに入ってたか
         for num in range(4):
            for num2 in range(cycle_number):
               if num2 not in p_set[player[num]]:
                  if player[num] in senior:
                     #先輩は連サイなるべくなくす
                     cycle_people[num2+1] = cycle_people[num2+1] + 3
                  else:               
                     cycle_people[num2+1] = cycle_people[num2+1] + 1

         #各サイクルに入る確率と連サイ数をリセット
         rate = [0 for num in range(cycle_number)]
         cycle_continuous = [0 for num in range(cycle_number)]

         #各サイクルに入る確率を決定
         for num in remain_cycle_all:
            #昼休憩がある場合は4,5連サイなし
            if num == afternoon-1 and lunch_break == 1:
               cycle_continuous[num] = cycle_people[num]
            elif num == afternoon and lunch_break == 1:
               cycle_continuous[num] = cycle_people[num+2]
            else:
               cycle_continuous[num] = cycle_people[num] + cycle_people[num+2]
            #連サイの数だけ確率を下げる
            rate[num] = pow(10,50) / pow((cycle_continuous[num] + 1),7)       
            #試合係が3人より多くなるの場合はrate=0
            if match_engagement_exist > 0 and match_engagement_number[num] + match_engagement_exist > 3:
               rate[num] = 0
            #限界サイクル数以上は入れないようにする
            if n < male_game_number and len(player_c[num]) >= male_game_number_percycle:
               rate[num] = 0
            #試合が少ないサイクルに優先的に入れるようにする
            if n < male_game_number:
               #男子の場合
               rate[num] = rate[num] * pow((male_game_number_percycle-len(player_c[num])),4)
            else:
               #女子は1サイクル目と午後はじめは優先的に試合を埋めるようにする
               if num == 0 or (num == afternoon and lunch_break == 1):
                  rate[num] = rate[num] * pow((game_number_percycle+3-len(player_c[num])),5)#何乗の数値は要調整
               else:
                  rate[num] = rate[num] * pow((game_number_percycle-len(player_c[num])),5)
            #入れてほしくないサイクル
            for num2 in range(not_specified_number):
               for num3 in range(4):
                  if player[num3] == cycle_not_specified[0][num2] and num ==  cycle_not_specified[1][num2]:
                     rate[num] = 0
            

         #3連サイチェック
         if three_check == 1:
            three_consecutive = [0 for num in range(cycle_number)]      
            for num in range(4):
               for num2 in range(cycle_number):
                  #前にも後にも入っている場合
                  if num2 != afternoon and num2 != afternoon-1:
                     if num2-1 in enter_cycle[num] and num2+1 in enter_cycle[num]:
                        three_consecutive[num2] += 1
                  #後ろ2つに入っている場合
                  if num2 != afternoon -2 and num2 != afternoon -1:
                     if num2+1 in enter_cycle[num] and num2+2 in enter_cycle[num]:
                        three_consecutive[num2] += 1
                  #前2つに入っている場合
                  if num2 != afternoon  and num2 != afternoon +1:
                     if num2-1 in enter_cycle[num] and num2-2 in enter_cycle[num]:
                        three_consecutive[num2] += 1            
            #3連サイを完全になしにする場合
            for num in range(cycle_number):
               if three_consecutive[num] >= 1:
                  rate[num] = 0
           
         #班長入場に出る人は最終サイクルに入れない
         for num in range(4):
            if player[num] in leader_admission:
               rate[cycle_number - 1] = 0
         #庶務チーフと試合チーフは最終サイクルに入れない
         for num in range(4):
            if player[num] in shomu_chief or player[num] in shiai_chief:
               rate[cycle_number - 1] = 0
         #昼休憩ありの場合は会計は昼前、昼後には入れない。試合チーフは昼前には入れない
         if lunch_break == 1:
            for num in range(4):
               if player[num] in treasurer:
                  rate[afternoon -1] = 0
                  rate[afternoon] = 0
               if player[num] in shiai_chief:
                  rate[afternoon -1] = 0
         #先輩は第1サイクルにはに入れない
         for num in range(4):
            if player[num] in senior:
               rate[0] = 0
         #1サイはコート数より多くならないように
         if len(player_c[0]) >= court_number:
            rate[0] = 0
         #昼休憩ありの場合は午後はじめもコート数より多くならないように
         if len(player_c[afternoon]) >= court_number and lunch_break == 1:
            rate[afternoon] = 0


         #各サイクルの確率を足し合わせる      
         sum_rate = 0
         for num in range(cycle_number):
            sum_rate = sum_rate + rate[num]
                    
         if sum_rate != 0:
            
            #どのサイクルに入るかを乱数で決定
            random_number = random.randrange(sum_rate)
            #下のfor文に使うためにリセット 
            sum_rate = 0
              
            for num in range(cycle_number):
               if sum_rate <= random_number and random_number < sum_rate + rate[num]:
                  dec_cycle = num
                  continuous_cycle += cycle_continuous[num]
               sum_rate = sum_rate + rate[num]
            
            #枠を消す
            b_list.remove(dec_cycle)
                
            #まだ選手が入っていないサイクルから消す
            for num in range(4):
               p_set[player[num]].remove(dec_cycle)
                 
            #サイクルに選手を追加
            player_c[dec_cycle].append(player_g[n])

            #試合係が入ったらカウント
            match_engagement_number[dec_cycle] += match_engagement_exist

      
                 
         else:
            ng_timetable = 1
            break

   #1サイはコート数ぴったりになるように
   if len(player_c[0]) != court_number:
      ng_timetable = 1
   #昼休憩ありの場合は午後はじめもコート数よぴったりになるように
   if len(player_c[afternoon]) != court_number and lunch_break == 1:
      ng_timetable = 1

      

   if ng_timetable == 0:
      #合計連サイ数
      sum_continuous_cycle += continuous_cycle

      #連サイ記録更新なるか…
      #一番低い記録を求める
      maximum = 0
      for num in range(10):
         if best_continuous_cycle[maximum] < best_continuous_cycle[num]:
            maximum = num
               
      #一番低い記録を更新する
      if best_continuous_cycle[maximum] > continuous_cycle:
         best_continuous_cycle[maximum] = continuous_cycle
         best_timetable[maximum] = [player_c[loop] for loop in range(cycle_number)]

            
   else:
      ng_number += 1


print("タイムテーブル作成完了")
for num in range(10):
   print("sheet"+str(num+1)+":連続サイクル数"+str(best_continuous_cycle[num]))
print("有効タイムテーブル率 "+str(loop_number-ng_number)+"/"+str(loop_number))
print("平均連サイ数"+str(round(sum_continuous_cycle/(loop_number-ng_number),1)))

wb2 = openpyxl.load_workbook("timetable_original.xlsx")
sheet1 = wb2["Sheet1"]
sheet2 = wb2["Sheet2"]
sheet3 = wb2["Sheet3"]
sheet4 = wb2["Sheet4"]
sheet5 = wb2["Sheet5"]
sheet6 = wb2["Sheet6"]
sheet7 = wb2["Sheet7"]
sheet8 = wb2["Sheet8"]
sheet9 = wb2["Sheet9"]
sheet10 = wb2["Sheet10"]

#試合に入った回数ごとに仕分け
count_list = []

for count in range(5):
   count_list.append([])
   count_list[count] = [name[num] for num in range(name_number) if player_count[num] == count+1]


#タイムテーブル書き込み開始
for top in range(10):
   g_number = 1
   #エクセルに書き込む
   for num in range(cycle_number):
      for num2 in range(len(best_timetable[top][num])):            
         exec("sheet%d.cell(num*4+4,num2*2+3,name[best_timetable[top][num][num2][0]])"%(top+1))
         exec("sheet%d.cell(num*4+5,num2*2+3,name[best_timetable[top][num][num2][1]])"%(top+1))
         exec("sheet%d.cell(num*4+4,num2*2+4,name[best_timetable[top][num][num2][2]])"%(top+1))
         exec("sheet%d.cell(num*4+5,num2*2+4,name[best_timetable[top][num][num2][3]])"%(top+1))
         exec("sheet%d.cell(num*4+6,num2*2+3,g_number)"%(top+1))
         g_number += 1
   
   if kainai == 1:
      #班名入力
      for num in range(cycle_number):
         for num2 in range(len(best_timetable[top][num])):
            for num3 in range(6):
               if best_timetable[top][num][num2][0] in group_structure[num3] and best_timetable[top][num][num2][1] in group_structure[num3]:
                  exec("sheet%d.cell(num*4+3,num2*2+3,group_name[num3])"%(top+1))
               if best_timetable[top][num][num2][2] in group_structure[num3] and best_timetable[top][num][num2][3] in group_structure[num3]:
                  exec("sheet%d.cell(num*4+3,num2*2+4,group_name[num3])"%(top+1))
      #各サイクルにその班が何回入ったか表示
      for num in range(6):
         exec("sheet%d.cell(44,13+num,group_name[num])"%(top+1))
                  
   #連サイ数を書き込む
   rensai = "連サイ数："+str(best_continuous_cycle[top])
   exec("sheet%d.cell(1,2,rensai)"%(top+1))

   #昼休憩and午後何サイからか
   exec("sheet%d.cell(44,10,lunch)"%(top+1))
   exec("sheet%d.cell(45,10,afternoon_time)"%(top+1))
   exec("sheet%d.cell(44,10).font = Font(size = 12)"%(top+1))
   exec("sheet%d.cell(45,10).font = Font(size = 12)"%(top+1))

   #試合係を表示
   for num in range(len(match_engagement_name)):
       exec("sheet%d.cell(45+num,8,match_engagement_name[num])"%(top+1))

   
   
   #試合に入った回数とサイクル数を記録
   sum_people = "合計"
   exec("sheet%d.cell(2,34,sum_people)"%(top+1))
   nin = "人"
   exec("sheet%d.cell(2,35,str(name_number)+ nin)"%(top+1))

   kai = "回"
   over50 = 0
   for count in range(5):
      exec("style = sheet%d.cell(4,(count + over50)*2+34)"%(top+1))
      style.value = str(count+1)+kai
      style.fill = PatternFill(patternType='solid', fgColor='FFFFFF00')
      for num in range(len(count_list[count])):
         exec("sheet%d.cell(num+5-math.floor(num/50)*50,(count+ over50 + math.floor(num/50))*2+34,count_list[count][num])"%(top+1))
      over50 += math.floor(len(count_list[count])/50)
         
for num in range(100):
   file_name = "timetable[" + str(num) + "].xlsx"
   if os.path.exists(file_name) == False:
      wb2 .save(file_name)
      print("timetable[" + str(num) + "].xlxsに保存しました")
      break
else:
   print("保存できませんでした(ファイル数が多すぎます)")